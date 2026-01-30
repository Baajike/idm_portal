from django.core.management.base import BaseCommand
from users.models import Staff
from openpyxl import load_workbook
from django.conf import settings
import os


class Command(BaseCommand):
    help = "Import staff from Excel file"

    def handle(self, *args, **kwargs):
        file_path = os.path.join(settings.BASE_DIR, "staff.xlsx")

        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR("Excel file not found"))
            return

        wb = load_workbook(file_path)
        sheet = wb.active

        created = 0
        skipped = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            name, functional_area, staff_id, contact, photo_name = row[:5]

            if not contact:
                skipped += 1
                continue

            staff, created_flag = Staff.objects.get_or_create(

                contact=contact,
                defaults={
                    "name": name,
                    "functional_area": functional_area,
                    "staff_id": staff_id,
                }
            )

            if photo_name:
                photo_path = f"staff_photos/{photo_name}"
                full_path = os.path.join(settings.MEDIA_ROOT, photo_path)

                if os.path.exists(full_path):
                    staff.photo = photo_path
                    staff.save()

            if created_flag:
                created += 1
            else:
                skipped += 1

        self.stdout.write(self.style.SUCCESS(
            f"Import complete. Created: {created}, Skipped: {skipped}"
        ))
